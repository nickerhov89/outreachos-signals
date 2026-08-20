"""OutreachOS Signals — FastAPI web app.
Reads from data/signals.db, serves dashboard, niche pages, signal detail, leads, XLSX export.
Run: uvicorn web.app:app --host 0.0.0.0 --port 8000"""
import io
import json
import sqlite3
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "signals.db"

app = FastAPI(
    title="OutreachOS Signals",
    description="Signal-as-a-Service for B2B outbound. 10 niches, real-time.",
    version="0.1.0",
)

# Static + templates
app.mount("/static", StaticFiles(directory=str(ROOT / "web" / "static")), name="static")
templates = Jinja2Templates(directory=str(ROOT / "web" / "templates"))


# Niche metadata (10 niches)
NICHES = [
    {"id": 1, "name": "CRM / BPM / Automation", "emoji": "🔄", "color": "#3b82f6"},
    {"id": 2, "name": "HRTech / Recruitment", "emoji": "👥", "color": "#8b5cf6"},
    {"id": 3, "name": "MarTech / Analytics", "emoji": "📊", "color": "#ec4899"},
    {"id": 4, "name": "AI / Data", "emoji": "🤖", "color": "#10b981"},
    {"id": 5, "name": "B2B SaaS", "emoji": "💼", "color": "#f59e0b"},
    {"id": 6, "name": "IT Services / Integration", "emoji": "🔧", "color": "#06b6d4"},
    {"id": 7, "name": "Cloud / Infrastructure", "emoji": "☁️", "color": "#6366f1"},
    {"id": 8, "name": "FinTech / Payments", "emoji": "💳", "color": "#84cc16"},
    {"id": 9, "name": "EdTech / LMS", "emoji": "🎓", "color": "#f97316"},
    {"id": 10, "name": "Marketplace / E-commerce", "emoji": "🛒", "color": "#ec4899"},
]
NICHES_BY_ID = {n["id"]: n for n in NICHES}

SOURCE_COLORS = {
    "greenhouse": "#22c55e",
    "lever": "#3b82f6",
    "apify_threads": "#a855f7",
    "funding_news": "#f59e0b",
    "github": "#1f2937",
    "linkedin_change": "#0a66c2",
    "reddit": "#ff4500",
    "hackernews": "#ff6600",
    "podcast": "#ec4899",
    "g2": "#ff492c",
    "job_text_tech": "#06b6d4",
}


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _fmt_age(iso_date: str | None) -> str:
    """Format '2026-08-19T00:00:00' as '2 days ago'."""
    if not iso_date:
        return "—"
    try:
        if "T" in iso_date:
            dt = datetime.fromisoformat(iso_date.replace("Z", "+00:00"))
        else:
            dt = datetime.strptime(iso_date, "%Y-%m-%d")
        # Naive: treat as UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        delta = datetime.now(timezone.utc) - dt
        days = delta.days
        if days == 0:
            return "today"
        if days == 1:
            return "1d ago"
        if days < 7:
            return f"{days}d ago"
        if days < 30:
            return f"{days // 7}w ago"
        return f"{days // 30}mo ago"
    except Exception:
        return iso_date[:10] if iso_date else "—"


# Add to Jinja env
templates.env.filters["age"] = _fmt_age


def _fromjson(s: str | None):
    if not s:
        return []
    try:
        return json.loads(s)
    except Exception:
        return []


templates.env.filters["fromjson"] = _fromjson


def _get_total_events() -> int:
    try:
        with get_db() as c:
            return c.execute("SELECT COUNT(*) FROM signal_events").fetchone()[0]
    except Exception:
        return 0


# Make available to all templates
templates.env.globals["range_niches"] = NICHES
templates.env.globals["total_events"] = _get_total_events


# ─── ROUTES ───────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    """Main dashboard: KPIs + recent signals by niche."""
    with get_db() as c:
        kpis = {
            "total_events": c.execute("SELECT COUNT(*) FROM signal_events").fetchone()[0],
            "total_classified": c.execute("SELECT COUNT(*) FROM signal_classifications").fetchone()[0],
            "total_with_contacts": c.execute(
                "SELECT COUNT(*) FROM signal_classifications WHERE buyer_contacts_json IS NOT NULL"
            ).fetchone()[0],
            "events_24h": c.execute(
                "SELECT COUNT(*) FROM signal_events WHERE collected_at > datetime('now', '-1 day')"
            ).fetchone()[0],
        }
        # Per-niche counts (with classification if available)
        niche_stats = []
        for n in NICHES:
            row = c.execute("""
                SELECT COUNT(DISTINCT cl.event_id) as classified
                FROM signal_classifications cl
                WHERE cl.niche_id = ?
            """, (n["id"],)).fetchone()
            events = c.execute("""
                SELECT COUNT(DISTINCT e.event_id)
                FROM signal_events e
                LEFT JOIN signal_classifications cl ON cl.event_id = e.event_id
                WHERE cl.niche_id = ? OR (cl.niche_id IS NULL AND e.raw_text LIKE ?)
            """, (n["id"], f"%{n['name'].split()[0]}%")).fetchone()[0]
            niche_stats.append({
                **n,
                "event_count": events,
                "classified_count": row["classified"],
            })
        # Top sources
        sources = c.execute("""
            SELECT source, COUNT(*) as c
            FROM signal_events
            GROUP BY source
            ORDER BY c DESC
        """).fetchall()
        # Recent 10 high-score signals
        recent = c.execute("""
            SELECT e.event_id, e.source, e.event_type, e.company_name, e.raw_text,
                   e.evidence_url, e.collected_at, cl.score, cl.first_angle, cl.niche_id
            FROM signal_events e
            JOIN signal_classifications cl ON cl.event_id = e.event_id
            WHERE cl.score >= 5
            ORDER BY cl.score DESC, e.collected_at DESC
            LIMIT 10
        """).fetchall()
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "kpis": kpis,
        "niche_stats": niche_stats,
        "sources": [dict(r) for r in sources],
        "recent": [dict(r) for r in recent],
        "source_colors": SOURCE_COLORS,
    })


@app.get("/niche/{niche_id}", response_class=HTMLResponse)
async def niche_page(request: Request, niche_id: int, min_score: float = 0):
    """Signals for a specific niche."""
    if niche_id not in NICHES_BY_ID:
        raise HTTPException(404, f"Unknown niche {niche_id}")
    niche = NICHES_BY_ID[niche_id]
    with get_db() as c:
        if min_score > 0:
            rows = c.execute("""
                SELECT e.event_id, e.source, e.event_type, e.company_name, e.raw_text,
                       e.evidence_url, e.evidence_snippet, e.collected_at,
                       cl.score, cl.first_angle, cl.niche_id, cl.buyer_contacts_json
                FROM signal_events e
                JOIN signal_classifications cl ON cl.event_id = e.event_id
                WHERE cl.niche_id = ? AND cl.score >= ?
                ORDER BY cl.score DESC, e.collected_at DESC
                LIMIT 200
            """, (niche_id, min_score)).fetchall()
        else:
            rows = c.execute("""
                SELECT e.event_id, e.source, e.event_type, e.company_name, e.raw_text,
                       e.evidence_url, e.evidence_snippet, e.collected_at,
                       cl.score, cl.first_angle, cl.niche_id, cl.buyer_contacts_json
                FROM signal_events e
                LEFT JOIN signal_classifications cl ON cl.event_id = e.event_id
                WHERE (cl.niche_id = ? OR e.raw_text LIKE ?)
                ORDER BY e.collected_at DESC
                LIMIT 200
            """, (niche_id, f"%{niche['name'].split()[0]}%")).fetchall()
    return templates.TemplateResponse("niche.html", {
        "request": request,
        "niche": niche,
        "signals": [dict(r) for r in rows],
        "min_score": min_score,
        "source_colors": SOURCE_COLORS,
    })


@app.get("/signal/{event_id}", response_class=HTMLResponse)
async def signal_detail(request: Request, event_id: str):
    """Single signal detail with all metadata + buyer contacts."""
    with get_db() as c:
        row = c.execute("""
            SELECT e.*, cl.score, cl.niche_id, cl.first_angle, cl.case_match,
                   cl.buyer_contacts_json, cl.icp_match, cl.evidence_strength, cl.urgency
            FROM signal_events e
            LEFT JOIN signal_classifications cl ON cl.event_id = e.event_id
            WHERE e.event_id = ?
        """, (event_id,)).fetchone()
    if not row:
        raise HTTPException(404, f"Unknown event {event_id}")
    sig = dict(row)
    if sig.get("buyer_contacts_json"):
        try:
            sig["buyer_contacts"] = json.loads(sig["buyer_contacts_json"])
        except Exception:
            sig["buyer_contacts"] = []
    else:
        sig["buyer_contacts"] = []
    if sig.get("raw_metadata"):
        try:
            sig["raw_metadata_obj"] = json.loads(sig["raw_metadata"])
        except Exception:
            sig["raw_metadata_obj"] = {}
    return templates.TemplateResponse("signal.html", {
        "request": request,
        "signal": sig,
        "source_colors": SOURCE_COLORS,
    })


@app.get("/leads", response_class=HTMLResponse)
async def leads_page(request: Request, min_score: float = 7.0):
    """Top leads: classified signals with score >= threshold."""
    with get_db() as c:
        rows = c.execute("""
            SELECT e.event_id, e.source, e.company_name, e.raw_text, e.collected_at,
                   cl.score, cl.first_angle, cl.case_match, cl.buyer_contacts_json
            FROM signal_events e
            JOIN signal_classifications cl ON cl.event_id = e.event_id
            WHERE cl.score >= ?
            ORDER BY cl.score DESC, e.collected_at DESC
            LIMIT 100
        """, (min_score,)).fetchall()
    leads = []
    for r in rows:
        d = dict(r)
        if d.get("buyer_contacts_json"):
            try:
                d["buyer_contacts"] = json.loads(d["buyer_contacts_json"])
            except Exception:
                d["buyer_contacts"] = []
        else:
            d["buyer_contacts"] = []
        leads.append(d)
    return templates.TemplateResponse("leads.html", {
        "request": request,
        "leads": leads,
        "min_score": min_score,
        "source_colors": SOURCE_COLORS,
    })


@app.get("/api/signals")
async def api_signals(
    niche: int | None = None,
    min_score: float = 0,
    source: str | None = None,
    limit: int = 100,
):
    """JSON API: list signals, filterable by niche, score, source."""
    query = """
        SELECT e.event_id, e.source, e.event_type, e.company_name, e.company_domain,
               e.raw_text, e.evidence_url, e.collected_at, e.event_date,
               cl.score, cl.niche_id, cl.first_angle, cl.case_match
        FROM signal_events e
        LEFT JOIN signal_classifications cl ON cl.event_id = e.event_id
        WHERE 1=1
    """
    params: list[Any] = []
    if niche is not None:
        query += " AND cl.niche_id = ?"
        params.append(niche)
    if min_score > 0:
        query += " AND (cl.score >= ? OR cl.score IS NULL AND 0 >= ?)"
        params.extend([min_score, min_score])
    if source:
        query += " AND e.source = ?"
        params.append(source)
    query += " ORDER BY cl.score DESC NULLS LAST, e.collected_at DESC LIMIT ?"
    params.append(min(limit, 1000))
    with get_db() as c:
        rows = c.execute(query, params).fetchall()
    return JSONResponse([dict(r) for r in rows])


@app.get("/api/leads")
async def api_leads(min_score: float = 7.0, limit: int = 100):
    """JSON API: top leads with contacts."""
    with get_db() as c:
        rows = c.execute("""
            SELECT e.event_id, e.company_name, e.company_domain, e.raw_text, e.evidence_url,
                   e.source, e.collected_at,
                   cl.score, cl.niche_id, cl.first_angle, cl.case_match, cl.buyer_contacts_json
            FROM signal_events e
            JOIN signal_classifications cl ON cl.event_id = e.event_id
            WHERE cl.score >= ?
            ORDER BY cl.score DESC
            LIMIT ?
        """, (min_score, min(limit, 1000))).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        if d.get("buyer_contacts_json"):
            try:
                d["buyer_contacts"] = json.loads(d.pop("buyer_contacts_json"))
            except Exception:
                d["buyer_contacts"] = []
        out.append(d)
    return JSONResponse(out)


@app.get("/api/export.xlsx")
async def export_xlsx(niche: int | None = None, min_score: float = 0):
    """Export leads to XLSX (requires openpyxl at runtime, else CSV fallback)."""
    query = """
        SELECT e.company_name, e.company_domain, e.source, e.event_type, e.raw_text,
               e.evidence_url, e.event_date, e.collected_at,
               cl.niche_id, cl.score, cl.icp_match, cl.evidence_strength, cl.urgency,
               cl.first_angle, cl.case_match, cl.buyer_contacts_json
        FROM signal_events e
        LEFT JOIN signal_classifications cl ON cl.event_id = e.event_id
        WHERE 1=1
    """
    params: list[Any] = []
    if niche is not None:
        query += " AND cl.niche_id = ?"
        params.append(niche)
    if min_score > 0:
        query += " AND (cl.score >= ?)"
        params.append(min_score)
    query += " ORDER BY cl.score DESC NULLS LAST LIMIT 5000"
    with get_db() as c:
        rows = [dict(r) for r in c.execute(query, params).fetchall()]

    # Try openpyxl, fall back to CSV
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Signals"
        headers = ["company", "domain", "source", "event_type", "signal", "evidence_url",
                   "event_date", "collected_at", "niche", "score", "icp_match", "evidence",
                   "urgency", "first_angle", "case_match", "contacts"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            contacts = r.pop("buyer_contacts_json", "")
            if contacts and contacts != "[]":
                try:
                    c_list = json.loads(contacts)
                    contacts_str = "; ".join(
                        f"{c.get('role','?')}: {c.get('email','?')} ({c.get('status','?')})"
                        for c in c_list
                    )
                except Exception:
                    contacts_str = str(contacts)
            else:
                contacts_str = ""
            r["contacts"] = contacts_str
            r.pop("buyer_contacts_json", None)
            ws.append([str(r.get(h, "")) for h in headers])
        # Column widths
        for col, w in zip("ABCDEFGHIJKLMNOP", [22, 25, 14, 14, 50, 35, 12, 18, 8, 8, 8, 8, 8, 50, 18, 50]):
            ws.column_dimensions[col].width = w
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=outreachos_signals.xlsx"},
        )
    except ImportError:
        # CSV fallback
        import csv
        buf = io.StringIO()
        if rows:
            w = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=outreachos_signals.csv"},
        )


@app.get("/health")
async def health():
    with get_db() as c:
        try:
            total = c.execute("SELECT COUNT(*) FROM signal_events").fetchone()[0]
            classified = c.execute("SELECT COUNT(*) FROM signal_classifications").fetchone()[0]
        except Exception as e:
            return JSONResponse({"status": "error", "error": str(e)}, status_code=500)
    return {
        "status": "ok",
        "events": total,
        "classified": classified,
        "ts": datetime.now(timezone.utc).isoformat(),
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("web.app:app", host="0.0.0.0", port=8000, reload=False)
